import argparse
import datetime as dt
import json
import os
import sqlite3
from pathlib import Path

import duckdb
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent
MODEL_STATE_PATH = ROOT / 'model_state.json'
DUCKDB_PATH = ROOT / 'mlb_quant_dashboard.duckdb'
SPORTSBOOK_DB_PATH = ROOT / 'sportsbook_lines.db'
WORKBOOK_PATH = ROOT / 'MLB_Season_To_Date_2026.xlsx'
REPORT_GLOB = 'MLB_Report_{date}.txt'
WATCHLIST_GLOB = 'Sportsbook_Watchlist_{date}.txt'
OUTPUT_GLOB = 'OPENING_DAY_READINESS_{date}.txt'

EXPECTED_SHEETS = ['Team_Metrics', 'Top_20_Batters', 'Top_20_Pitchers', 'Power_Rankings']
EXPECTED_STATE_BLOCKS = [
    'probability_shrinkage',
    'total_shrinkage',
    'margin_shrinkage',
    'lineup_earn_back',
    'realized_uncertainty',
    'total_sigma_calibration',
]


class CheckLog:
    def __init__(self):
        self.rows = []

    def add(self, status: str, category: str, name: str, detail: str):
        self.rows.append({'status': status, 'category': category, 'name': name, 'detail': detail})

    def counts(self):
        out = {'PASS': 0, 'WARN': 0, 'FAIL': 0}
        for row in self.rows:
            out[row['status']] = out.get(row['status'], 0) + 1
        return out


def load_json(path: Path):
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        return {}


def check_files(log: CheckLog, report_date: str):
    report_path = ROOT / REPORT_GLOB.format(date=report_date)
    watchlist_path = ROOT / WATCHLIST_GLOB.format(date=report_date)
    for path, label, severity in [
        (ROOT / 'run_daily_mlb_report.py', 'Primary runner', 'FAIL'),
        (ROOT / 'quant_dashboard.py', 'Quant dashboard', 'FAIL'),
        (WORKBOOK_PATH, 'Season workbook', 'FAIL'),
        (DUCKDB_PATH, 'Quant DuckDB store', 'FAIL'),
        (SPORTSBOOK_DB_PATH, 'Sportsbook DB', 'FAIL'),
        (MODEL_STATE_PATH, 'Model state', 'FAIL'),
        (report_path, 'Daily report', 'FAIL'),
        (watchlist_path, 'Daily watchlist', 'WARN'),
    ]:
        if path.exists():
            log.add('PASS', 'files', label, str(path))
        else:
            log.add(severity, 'files', label, f'Missing: {path.name}')
    return report_path, watchlist_path


def check_model_state(log: CheckLog):
    state = load_json(MODEL_STATE_PATH)
    if not state:
        log.add('FAIL', 'state', 'Model state JSON', 'Could not load model_state.json')
        return {}
    for key in EXPECTED_STATE_BLOCKS:
        if key in state:
            log.add('PASS', 'state', f'State block `{key}`', 'Present')
        else:
            log.add('FAIL', 'state', f'State block `{key}`', 'Missing critical state block')
    return state


def check_workbook(log: CheckLog):
    if not WORKBOOK_PATH.exists():
        return
    try:
        wb = load_workbook(WORKBOOK_PATH, read_only=True)
    except Exception as exc:
        log.add('FAIL', 'workbook', 'Workbook load', f'Could not open workbook: {exc}')
        return
    sheetnames = wb.sheetnames
    for sheet in EXPECTED_SHEETS:
        if sheet in sheetnames:
            rows = wb[sheet].max_row
            status = 'PASS' if rows and rows > 1 else 'WARN'
            log.add(status, 'workbook', f'Sheet `{sheet}`', f'Rows: {rows}')
        else:
            log.add('FAIL', 'workbook', f'Sheet `{sheet}`', 'Missing expected sheet')


def check_sportsbook_db(log: CheckLog):
    if not SPORTSBOOK_DB_PATH.exists():
        return
    try:
        with sqlite3.connect(SPORTSBOOK_DB_PATH) as conn:
            tables = {row[0] for row in conn.execute("SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
            for table in ['bet_journal', 'odds_snapshots']:
                if table in tables:
                    count = conn.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
                    status = 'PASS' if count > 0 else 'WARN'
                    log.add(status, 'market-db', f'Table `{table}`', f'Rows: {count}')
                else:
                    log.add('FAIL', 'market-db', f'Table `{table}`', 'Missing expected table')
    except Exception as exc:
        log.add('FAIL', 'market-db', 'Sportsbook DB access', str(exc))


def check_duckdb(log: CheckLog, report_date: str):
    if not DUCKDB_PATH.exists():
        return
    try:
        con = duckdb.connect(str(DUCKDB_PATH), read_only=True)
    except Exception as exc:
        log.add('FAIL', 'duckdb', 'DuckDB open', str(exc))
        return
    required_tables = [
        'quant_runs',
        'prediction_snapshots',
        'policy_research_recommendations',
        'policy_research_market_type_recommendations',
        'market_proof_summary',
        'market_proof_by_market',
    ]
    for table in required_tables:
        try:
            count = con.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
            status = 'PASS' if (count > 0 or table in {'market_proof_by_market'}) else 'WARN'
            log.add(status, 'duckdb', f'Table `{table}`', f'Rows: {count}')
        except Exception as exc:
            log.add('FAIL', 'duckdb', f'Table `{table}`', str(exc))
    try:
        latest = con.execute(
            """
            SELECT report_date, created_ts, live_validation_games, live_validation_settled_days,
                   market_policy_alignment_status, market_policy_moneyline_policy,
                   market_policy_total_policy, market_policy_run_line_policy,
                   market_proof_total_bets
            FROM quant_runs
            ORDER BY created_ts DESC
            LIMIT 1
            """
        ).fetchone()
        if latest:
            latest_report_date = str(latest[0])
            status = 'PASS' if latest_report_date == report_date else 'FAIL'
            log.add(status, 'duckdb', 'Latest run date', f'Latest run report_date={latest_report_date}, expected={report_date}')
            log.add('PASS', 'duckdb', 'Latest market policy state', f'ML={latest[5]} | Total={latest[6]} | RL={latest[7]}')
            if latest[8] in (None, 0):
                log.add('WARN', 'duckdb', 'Totals market proof sample', 'No settled totals priced bets yet')
            else:
                log.add('PASS', 'duckdb', 'Totals market proof sample', f'Settled totals bets: {latest[8]}')
        else:
            log.add('FAIL', 'duckdb', 'Latest run date', 'No rows in quant_runs')
    except Exception as exc:
        log.add('FAIL', 'duckdb', 'Latest run inspection', str(exc))


def infer_go_no_go(counts):
    if counts.get('FAIL', 0) > 0:
        return 'NO-GO'
    if counts.get('WARN', 0) > 0:
        return 'GO WITH WARNINGS'
    return 'GO'


def build_text(log: CheckLog, report_date: str):
    counts = log.counts()
    status = infer_go_no_go(counts)
    lines = []
    lines.append(f'OPENING DAY READINESS - {report_date}')
    lines.append(f'Generated: {dt.datetime.now().isoformat()}')
    lines.append('')
    lines.append(f'Overall status: {status}')
    lines.append(f"PASS: {counts.get('PASS', 0)} | WARN: {counts.get('WARN', 0)} | FAIL: {counts.get('FAIL', 0)}")
    lines.append('')
    categories = []
    seen = set()
    for row in log.rows:
        if row['category'] not in seen:
            seen.add(row['category'])
            categories.append(row['category'])
    for category in categories:
        lines.append(category.upper())
        for row in [r for r in log.rows if r['category'] == category]:
            lines.append(f"[{row['status']}] {row['name']}: {row['detail']}")
        lines.append('')
    lines.append('Interpretation:')
    if status == 'NO-GO':
        lines.append('- One or more critical launch dependencies are broken or missing.')
    elif status == 'GO WITH WARNINGS':
        lines.append('- Core launch path is healthy, but some expected pre-regular-season gaps remain.')
    else:
        lines.append('- Core launch path is healthy with no current warnings.')
    return '\n'.join(lines)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--report-date', default=dt.date.today().isoformat())
    args = parser.parse_args()

    log = CheckLog()
    check_files(log, args.report_date)
    check_model_state(log)
    check_workbook(log)
    check_sportsbook_db(log)
    check_duckdb(log, args.report_date)

    text = build_text(log, args.report_date)
    out_path = ROOT / OUTPUT_GLOB.format(date=args.report_date)
    out_path.write_text(text, encoding='utf-8')
    print(out_path)
    print(text)


if __name__ == '__main__':
    main()
