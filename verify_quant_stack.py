from __future__ import annotations

import sqlite3

from openpyxl import load_workbook

import quant_store


def main() -> None:
    bundle = quant_store.load_dashboard_bundle()
    print('STATUS', bundle.get('status'))

    for key in [
        'run_history',
        'latest_run',
        'predictions',
        'candidates',
        'portfolio_recommendations',
        'portfolio_summary',
        'portfolio_market_exposure',
        'portfolio_team_exposure',
        'portfolio_game_exposure',
        'clean_backtest',
        'live_archive',
        'bet_journal',
        'odds_snapshots',
    ]:
        value = bundle.get(key)
        shape = getattr(value, 'shape', None)
        print(key.upper(), shape)

    try:
        with sqlite3.connect('sportsbook_lines.db') as conn:
            raw_odds = conn.execute('SELECT COUNT(*) FROM odds_snapshots').fetchone()[0]
            raw_journal = conn.execute('SELECT COUNT(*) FROM bet_journal').fetchone()[0]
        print('RAW_ODDS_SNAPSHOTS', raw_odds)
        print('RAW_BET_JOURNAL', raw_journal)
    except Exception as exc:
        print('RAW_SQLITE_CHECK_ERR', exc)

    latest_run = bundle.get('latest_run')
    if getattr(latest_run, 'empty', True) is False:
        row = latest_run.iloc[0]
        print('LATEST_REPORT_DATE', row.get('report_date'))
        print('LATEST_PHASE', row.get('phase'))
        print('PREDICTION_COUNT', row.get('prediction_count'))
        print('ACTIONABLE_COUNT', row.get('actionable_market_count'))

    portfolio_summary = bundle.get('portfolio_summary')
    if getattr(portfolio_summary, 'empty', True) is False:
        summary_row = portfolio_summary.iloc[0]
        print('PORTFOLIO_ALLOCATED_UNITS', summary_row.get('allocated_units'))
        print('PORTFOLIO_ALLOCATED_COUNT', summary_row.get('allocated_candidate_count'))

    workbook = load_workbook('MLB_Season_To_Date_2026.xlsx', read_only=True)
    print('SHEETS', workbook.sheetnames)
    print('TEAM_ROWS', workbook['Team_Metrics'].max_row)
    print('BATTER_ROWS', workbook['Top_20_Batters'].max_row)
    print('PITCHER_ROWS', workbook['Top_20_Pitchers'].max_row)
    print('POWER_ROWS', workbook['Power_Rankings'].max_row)


if __name__ == '__main__':
    main()
