import duckdb
from openpyxl import load_workbook
import quant_store

con = duckdb.connect(r'mlb_quant_dashboard.duckdb', read_only=True)
for table in [
    'quant_runs',
    'policy_research_recommendations',
    'policy_research_window_rankings',
    'policy_research_shadow_daily',
    'policy_research_shadow_summary',
    'policy_research_market_daily',
    'policy_research_market_policy_daily',
    'policy_research_market_policy_summary',
    'policy_research_market_type_recommendations',
]:
    try:
        count = con.execute(f'SELECT COUNT(*) FROM {table}').fetchone()[0]
        print(f'{table}: {count}')
    except Exception as exc:
        print(f'{table}: ERR {exc}')

bundle = quant_store.load_dashboard_bundle()
for key in [
    'policy_research_history',
    'policy_research_rankings',
    'policy_research_shadow_daily',
    'policy_research_shadow_summary',
    'policy_research_market_daily',
    'policy_research_market_policy_daily',
    'policy_research_market_policy_summary',
    'policy_research_market_type_history',
]:
    frame = bundle.get(key)
    print(f'{key}: {getattr(frame, "shape", None)}')

latest_run = bundle.get('latest_run')
if latest_run is not None and not latest_run.empty:
    cols = [
        'report_date',
        'policy_research_policy',
        'policy_research_posture',
        'policy_research_window_label',
        'policy_research_evidence_source',
        'policy_research_evidence_strength',
        'market_policy_alignment_status',
        'market_policy_real_market_count',
        'market_policy_change_count',
        'market_policy_moneyline_policy',
        'market_policy_total_policy',
        'market_policy_run_line_policy',
        'live_validation_games',
        'live_validation_settled_days',
        'live_validation_log_loss',
        'live_validation_accuracy',
        'uncertainty_moneyline_realized_bias',
        'uncertainty_total_realized_bias',
        'uncertainty_run_line_realized_bias',
        'uncertainty_moneyline_realized_confidence',
        'uncertainty_total_realized_confidence',
        'uncertainty_run_line_realized_confidence',
    ]
    available = [c for c in cols if c in latest_run.columns]
    print('latest_run_policy:', latest_run[available].head(1).to_dict('records'))

wb = load_workbook('MLB_Season_To_Date_2026.xlsx', read_only=True)
print('sheets:', wb.sheetnames)
print('rows:', wb['Team_Metrics'].max_row, wb['Top_20_Batters'].max_row, wb['Top_20_Pitchers'].max_row, wb['Power_Rankings'].max_row)






